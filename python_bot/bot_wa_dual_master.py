import os
import re
import shutil
import openpyxl
import glob
import requests
import calendar
import base64
from flask import Flask, request, jsonify
from datetime import datetime

from fpdf import FPDF
from PIL import Image, ImageDraw, ImageFont

app = Flask(__name__)

# ================= KONFIGURASI DARI DOCKER (.ENV) =================
ALLOWED_USERS = os.getenv("ALLOWED_USERS", "6282328672766").split(",")
WA_API_TOKEN = os.getenv("WA_API_TOKEN", "zylvemedia")

# Di Docker, localhost diganti dengan nama service dari docker-compose
WA_API_BOT = "http://baileys_bot:8000/send-message"       
WA_API_PELANGGAN = "http://baileys_pelanggan:8001/send-message" 

NAS_PATH = os.getenv("NAS_PATH", "/mnt/nas/share/bot_pelanggan/")
if not NAS_PATH.endswith('/'): NAS_PATH += '/'

TEMPLATE_MASTER = os.path.join(NAS_PATH, 'template-import-pelanggan.xlsx')
KTP_PATH = os.path.join(NAS_PATH, 'foto_ktp/')
RUMAH_PATH = os.path.join(NAS_PATH, 'foto_rumah/')
EXCEL_PATH = os.path.join(NAS_PATH, 'data_excel/')
PDF_PATH = os.path.join(NAS_PATH, 'data_pdf/')

for path in [KTP_PATH, RUMAH_PATH, EXCEL_PATH, PDF_PATH]:
    os.makedirs(path, exist_ok=True)

# State Management
user_data = {}
tag_data = {} 

QUESTIONS = [
    ('nama_pelanggan*', 'Nama Pelanggan *'),
    ('nik*', 'No KTP / NIK (Pastikan 16 Digit)'),
    ('no_hp*', 'No. WhatsApp (Hanya Angka, Contoh: 08123456xxx)'),
    ('alamat*', 'Alamat Lengkap'),
    ('tgl_aktif*', 'Tanggal Aktif Pemasangan (Format: DD/MM/YYYY, Contoh: 15/01/2026)'),
    ('redaman', 'Nilai Redaman FO (Contoh: -22 atau -24)'),
    ('tgl_jatuh_tempo', 'Angka Tanggal Jatuh Tempo Saja (Input: 1 sampai 31, Contoh: 15)'),
    ('biaya_paket', 'Biaya Paket Bulanan (Hanya Angka, Contoh: 150000)')
]

TAG_QUESTIONS = [
    ('tgl_aktif', 'Tanggal Aktif Pemasangan (Format: DD/MM/YYYY, Contoh: 15/01/2026)'),
    ('tgl_jatuh_tempo', 'Angka Tanggal Tagihan / Jatuh Tempo (Input: 1 sampai 31, Contoh: 15)'),
    ('biaya_paket', 'Nominal Paket Bulanan (Hanya Angka, Contoh: 150000)'),
    ('no_hp', 'No. WhatsApp Pelanggan (Hanya Angka, Contoh: 08123456xxx)')
]

def hitung_biaya_prorata(tgl_aktif_str, tgl_jt_day_str, biaya_paket_str):
    try:
        tgl_aktif = datetime.strptime(tgl_aktif_str, '%Y-%m-%d')
        day_jt = int(tgl_jt_day_str)
        biaya = int(biaya_paket_str)
        
        tahun = tgl_aktif.year
        bulan = tgl_aktif.month
        
        max_day = calendar.monthrange(tahun, bulan)[1]
        target_day = min(day_jt, max_day)
        tgl_jt_target = datetime(tahun, billing_month := bulan, target_day)
        
        if tgl_jt_target <= tgl_aktif:
            bulan += 1
            if bulan > 12:
                bulan = 1
                tahun += 1
            max_day = calendar.monthrange(tahun, bulan)[1]
            target_day = min(day_jt, max_day)
            tgl_jt_target = datetime(tahun, bulan, target_day)
        
        jumlah_hari = (tgl_jt_target - tgl_aktif).days - 1
        if jumlah_hari < 0:
            jumlah_hari = 0
            
        biaya_prorata = (biaya / 30) * jumlah_hari
        return int(biaya_prorata), jumlah_hari
    except Exception:
        return 0, 0

def beri_watermark_foto(image_path, lat, lon):
    try:
        img = Image.open(image_path)
        draw = ImageDraw.Draw(img)
        waktu_sekarang = datetime.now().strftime('%d/%m/%Y | %H:%M:%S')
        teks_watermark = f"📍 LOKASI : {lat}, {lon}\n🕒 WAKTU  : {waktu_sekarang} WIB\n🔒 ARSIP  : NAS SERVER AUTOMATION"
        
        width, height = img.size
        font_size = max(18, int(height * 0.025))
        
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
        except IOError:
            try:
                font = ImageFont.load_default(size=font_size)
            except Exception:
                font = ImageFont.load_default()
            
        text_box = draw.textbbox((0, 0), teks_watermark, font=font)
        text_w = text_box[2] - text_box[0]
        text_h = text_box[3] - text_box[1]
        
        x = 30
        y = height - text_h - 60
        
        draw.rectangle([x - 15, y - 15, x + text_w + 25, y + text_h + 25], fill="black")
        draw.text((x, y), teks_watermark, fill="white", font=font)
        img.save(image_path)
    except Exception as e:
        print(f"❌ [Watermark Error] Gagal cetak watermark: {str(e)}")

def send_wa_bot(target, text):
    payload = {'target': target, 'message': text}
    headers = {'Authorization': WA_API_TOKEN}
    try:
        r = requests.post(WA_API_BOT, json=payload, headers=headers, timeout=5)
        print(f"[API Outbound Bot] Tembak Balas ke {target} | Status HTTP: {r.status_code}")
    except Exception as e:
        print(f"❌ [API Outbound Error] Gagal kirim balik via port 8000: {str(e)}")

def kirim_wa_terima_kasih(no_hp, nama, tgl_jt, biaya_paket, tgl_aktif_str):
    if no_hp.startswith('0'):
        no_hp = '62' + no_hp[1:]
        
    biaya_prorata, jumlah_hari = hitung_biaya_prorata(tgl_aktif_str, tgl_jt, biaya_paket)
    
    try:
        prorata_format = f"{biaya_prorata:,}".replace(",", ".")
        normal_format = f"{int(biaya_paket):,}".replace(",", ".")
    except Exception:
        prorata_format = biaya_paket
        normal_format = biaya_paket
        
    pesan_teks = (
        f"Halo Bapak/Ibu *{nama}*,\n\n"
        f"Terima kasih telah memilih layanan internet kami. Sambungan Baru Anda saat ini telah *AKTIF*.\n\n"
        f"📌 *Informasi Pembayaran Pelanggan*:\n"
        f"• Nama Pelanggan: {nama}\n"
        f"• Tanggal Jatuh Tempo: Tiap Tanggal *{tgl_jt}* setiap bulannya.\n"
        f"• *Tagihan Bulan Ke-2 (Prorata {jumlah_hari} Hari)*: *Rp {prorata_format}*\n"
        f"  _(Dihitung proporsional dari hari ke-2 aktif hingga tanggal jatuh tempo terdekat)_\n"
        f"• Tagihan Bulan Ke-3 Normal: Rp {normal_format}\n\n"
        f"⚠️ *PENTING - HARAP SIMPAN NOMOR INI*:\n"
        f"Mohon simpan nomor WhatsApp Admin ini untuk mempermudah Anda menerima info tagihan bulanan serta mempercepat penanganan kendala teknis jika terjadi gangguan internet di rumah Anda.\n\n"
        f"Link Pembayaran : https://ibb.co.com/zTQNjd4c . Mohon kirim bukti pembayaran ke nomer ini, lakukan pembayaran tepat waktu sebelum tanggal jatuh tempo untuk menghindari isolir sistem otomatis. Terima kasih! 🙏"
    )
    
    payload = {'target': no_hp, 'message': pesan_teks}
    headers = {'Authorization': WA_API_TOKEN}
    
    try:
        response = requests.post(WA_API_PELANGGAN, json=payload, headers=headers, timeout=5)
        return response.status_code == 200, prorata_format, jumlah_hari
    except Exception as e:
        print(f"❌ [API Pelanggan Error] Gagal kirim nota via port 8001: {str(e)}")
        return False, prorata_format, jumlah_hari

def simpan_data_keseluruhan(chat_id):
    user_input = user_data[chat_id]['data']
    nama_asli = user_input.get('nama_pelanggan*', 'Tanpa_Nama')
    nama_file_interal = nama_asli.replace(" ", "_").replace("/", "_")
    
    lokasi_excel_baru = os.path.join(EXCEL_PATH, f"{nama_file_interal}.xlsx")
    lokasi_pdf_baru = os.path.join(PDF_PATH, f"BA_{nama_file_interal}.pdf")
    
    lat_val = user_input.get('lat*', '')
    lon_val = user_input.get('lon*', '')
    maps_link = f"http://maps.google.com/?q={lat_val},{lon_val}" if lat_val else "-"
    
    try:
        shutil.copy(TEMPLATE_MASTER, lokasi_excel_baru)
        wb = openpyxl.load_workbook(lokasi_excel_baru)
        ws = wb.active
        ws.cell(row=2, column=1).value = nama_asli
        ws.cell(row=2, column=4).value = user_input.get('no_hp*', '')
        ws.cell(row=2, column=5).value = user_input.get('alamat*', '')
        ws.cell(row=2, column=6).value = lat_val
        ws.cell(row=2, column=7).value = lon_val
        ws.cell(row=2, column=8).value = 'Non-PPN'
        ws.cell(row=2, column=9).value = user_input.get('nik*', '')
        ws.cell(row=2, column=12).value = user_input.get('tgl_aktif*', '')
        tgl_jt = user_input.get('tgl_jatuh_tempo', '-')
        redam = user_input.get('redaman', '-')
        ws.cell(row=2, column=22).value = f"Tgl Jatuh Tempo: Tiap Tanggal {tgl_jt} | Pengingat Isolir: AKTIF | Redaman: {redam} dBm | Maps: {maps_link}"
        wb.save(lokasi_excel_baru)
        wb.close()
    except Exception as e:
        print(f"❌ [Excel Error] {str(e)}")
    
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", 'B', 16)
        pdf.cell(0, 10, "BERITA ACARA PASANG BARU (BA-PSB)", ln=1, align='C')
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 5, f"NAS Server Log: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=1, align='C')
        pdf.ln(5); pdf.line(10, 28, 200, 28)
        
        pdf.set_font("Helvetica", 'B', 12); pdf.cell(0, 10, "DATA PELANGGAN", ln=1); pdf.set_font("Helvetica", size=11)
        pdf.cell(50, 8, "Nama Pelanggan"); pdf.cell(0, 8, f": {nama_asli}", ln=1)
        pdf.cell(50, 8, "No. NIK / KTP"); pdf.cell(0, 8, f": {user_input.get('nik*', '')}", ln=1)
        pdf.cell(50, 8, "No. WhatsApp"); pdf.cell(0, 8, f": {user_input.get('no_hp*', '')}", ln=1)
        pdf.cell(50, 8, "Alamat"); pdf.cell(0, 8, f": {user_input.get('alamat*', '')}", ln=1)
        
        tgl_aktif_pdf = user_input.get('tgl_aktif*', '-')
        if tgl_aktif_pdf and '-' in tgl_aktif_pdf:
            tgl_aktif_pdf = datetime.strptime(tgl_aktif_pdf, '%Y-%m-%d').strftime('%d/%m/%Y')
            
        pdf.cell(50, 8, "Tanggal Aktif"); pdf.cell(0, 8, f": {tgl_aktif_pdf}", ln=1)
        pdf.cell(50, 8, "Tanggal Jatuh Tempo"); pdf.cell(0, 8, f": Tiap Tanggal {tgl_jt}", ln=1)
        
        pdf.ln(5)
        pdf.set_font("Helvetica", 'B', 12); pdf.cell(0, 10, "DATA TEKNIS INSTALASI", ln=1); pdf.set_font("Helvetica", size=11)
        pdf.cell(50, 8, "Hasil Redaman FO"); pdf.cell(0, 8, f": {redam} dBm", ln=1)
        pdf.cell(50, 8, "Link Google Maps"); pdf.cell(0, 8, f": {maps_link}", ln=1)
        
        pdf.ln(15); pdf.cell(95, 8, "Teknisi Lapangan,", align='C'); pdf.cell(95, 8, "Pelanggan,", align='C', ln=1)
        pdf.ln(15); pdf.cell(95, 8, "( ____________________ )", align='C'); pdf.cell(95, 8, f"( {nama_asli} )", align='C', ln=1)
        
        pdf.add_page(); pdf.set_font("Helvetica", 'B', 14); pdf.cell(0, 10, "LAMPIRAN DOKUMENTASI (FOTO WATERMARK NAS)", ln=1, align='C'); pdf.ln(10)
        if 'final_ktp_path' in user_data[chat_id] and os.path.exists(user_data[chat_id]['final_ktp_path']):
            pdf.set_font("Helvetica", 'B', 11); pdf.cell(0, 8, "1. FOTO DOKUMEN KTP (LOKASI & JAM TER-REKAM)", ln=1)
            pdf.image(user_data[chat_id]['final_ktp_path'], x=15, y=pdf.get_y()+2, w=110); pdf.set_y(pdf.get_y() + 75)
        if 'final_rumah_path' in user_data[chat_id] and os.path.exists(user_data[chat_id]['final_rumah_path']):
            pdf.set_font("Helvetica", 'B', 11); pdf.cell(0, 8, "2. FOTO DOKUMENTASI RUMAH (LOKASI & JAM TER-REKAM)", ln=1)
            pdf.image(user_data[chat_id]['final_rumah_path'], x=15, y=pdf.get_y()+2, w=110)
            
        pdf.output(lokasi_pdf_baru)
    except Exception as e:
        print(f"❌ [PDF Error] {str(e)}")

    no_hp_target = user_input.get('no_hp*', '')
    biaya_paket = user_input.get('biaya_paket', '0')
    tgl_aktif_str = user_input.get('tgl_aktif*', '')
    
    wa_terkirim, prorata_print, hari_print = kirim_wa_terima_kasih(no_hp_target, nama_asli, tgl_jt, biaya_paket, tgl_aktif_str)
    status_wa = f"🚀 Notifikasi WA sukses dikirim via nomor utama (Prorata {hari_print} hari sebesar Rp {prorata_print})." if wa_terkirim else "⚠️ Gagal kirim WA otomatis."

    pesan_sukses = (
        f"✅ *PROSES PSB BERHASIL!*\n\nPelanggan *{nama_asli}* sukses didata.\n"
        f"• File Excel & PDF Berita Acara tersimpan aman di NAS.\n"
        f"• {status_wa}"
    )
    send_wa_bot(chat_id, pesan_sukses)
    del user_data[chat_id]

# ================= CORE ENGINE WEBHOOK =================
@app.route('/webhook', methods=['POST'])
def webhook_gateway():
    payload = request.json
    if not payload:
        return jsonify({"status": "error", "message": "Payload kosong"}), 400

    chat_id = payload.get('from', '') 
    body = payload.get('body', '').strip()
    msg_type = payload.get('type', 'chat')

    # Pemisahan Otomatis LID & JID agar tidak error
    sender_raw = chat_id.split('@')[0]
    is_lid = '@lid' in chat_id

    allowed_clean = []
    for u in ALLOWED_USERS:
        u_str = str(u).strip()
        if u_str.startswith('0'):
            u_str = '62' + u_str[1:]
        allowed_clean.append(u_str)

    sender_clean = sender_raw
    if not is_lid and sender_clean.startswith('0'):
        sender_clean = '62' + sender_clean[1:]

    print(f"\n📢 [CHAT MASUK] Dari: {sender_raw} (LID: {is_lid}) | Isi: {body}")

    if sender_clean not in allowed_clean:
        print(f"❌ [AKSES DITOLAK] Nomor/LID {sender_clean} tidak terdaftar di whitelist!")
        return jsonify({"status": "ignored"}), 200

    print(f"✅ [AKSES DIIZINKAN] Memproses...")

    if body.lower() in ['/batal', 'batal']:
        if chat_id in user_data:
            del user_data[chat_id]
            send_wa_bot(chat_id, "🚫 Sesi PSB dihentikan. Ketik *psb* atau */psb* untuk mulai baru.")
        elif chat_id in tag_data:
            del tag_data[chat_id]
            send_wa_bot(chat_id, "🚫 Sesi Penagihan dihentikan. Ketik *tag* atau */tag* untuk mulai baru.")
        else:
            send_wa_bot(chat_id, "Tidak ada sesi yang sedang berjalan.")
        return jsonify({"status": "success"}), 200

    if body.lower().startswith(('/cari', 'cari')):
        parts = body.split(' ', 1)
        if len(parts) > 1:
            keyword = parts[1].lower()
            files = [f for f in os.listdir(EXCEL_PATH) if f.endswith('.xlsx') and not f.startswith('~')]
            hasil = [f.replace('.xlsx', '') for f in files if keyword in f.lower()]
            pesan = f"✅ Ditemukan {len(hasil)} data:\n" + "\n".join([f"- {h}" for f in hasil]) if hasil else "❌ Data tidak ditemukan di NAS."
            send_wa_bot(chat_id, pesan)
        else:
            send_wa_bot(chat_id, "🔍 Format salah. Contoh: *cari Jhon*")
        return jsonify({"status": "success"}), 200

    if body.lower().startswith(('/hapus', 'hapus')):
        parts = body.split(' ', 1)
        if len(parts) > 1:
            nama_target = parts[1].replace(" ", "_").replace("/", "_")
            target_excel = os.path.join(EXCEL_PATH, f"{nama_target}.xlsx")
            target_pdf = os.path.join(PDF_PATH, f"BA_{nama_target}.pdf")
            
            if os.path.exists(target_excel):
                os.remove(target_excel)
                if os.path.exists(target_pdf): os.remove(target_pdf)
                for f in glob.glob(os.path.join(KTP_PATH, f"KTP_{nama_target}_*.jpg")) + glob.glob(os.path.join(RUMAH_PATH, f"RUMAH_{nama_target}_*.jpg")):
                    os.remove(f)
                send_wa_bot(chat_id, f"✅ Seluruh Data Excel, PDF, dan foto atas nama *{nama_target}* di-hapus dari NAS.")
            else:
                send_wa_bot(chat_id, f"❌ Data *{nama_target}* tidak ada.")
        else:
            send_wa_bot(chat_id, "🗑️ Format salah. Contoh: *hapus Jhon_Doe*")
        return jsonify({"status": "success"}), 200

    if body.lower() in ['/rekap', 'rekap']:
        bulan_ini = datetime.now().strftime('%Y-%m')
        files = [f for f in os.listdir(EXCEL_PATH) if f.endswith('.xlsx') and not f.startswith('~')]
        data_rekap = []
        
        send_wa_bot(chat_id, "⏳ Membaca seluruh file NAS...")
        for f in files:
            filepath = os.path.join(EXCEL_PATH, f)
            if datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m') == bulan_ini:
                try:
                    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                    ws = wb.active
                    tgl_aktif = ws.cell(row=2, column=12).value
                    tgl_aktif_str = str(tgl_aktif)[:10] if tgl_aktif else "-"
                    if '-' in tgl_aktif_str:
                        tgl_aktif_str = datetime.strptime(tgl_aktif_str, '%Y-%m-%d').strftime('%d/%m/%Y')
                    data_rekap.append(f"- {f.replace('.xlsx', '')} | Aktif: {tgl_aktif_str}")
                    wb.close()
                except Exception: pass
                    
        pesan = f"📊 *REKAP PELANGGAN BULAN INI*\nTotal: {len(data_rekap)} Pelanggan\n\n" + "\n".join(data_rekap) if data_rekap else "📊 Belum ada data pelanggan baru di bulan ini."
        send_wa_bot(chat_id, pesan)
        return jsonify({"status": "success"}), 200

    if body.lower() in ['/tag', 'tag'] and chat_id not in user_data and chat_id not in tag_data:
        tag_data[chat_id] = {'step': 0, 'data': {}}
        pesan = f"💳 *[TAGIHAN] Mulai Hitung Penagihan Bulan Ke-2*\n_Ketik batal untuk membatalkan._\n\nMasukkan *{TAG_QUESTIONS[0][1]}*:"
        send_wa_bot(chat_id, pesan)
        return jsonify({"status": "success"}), 200

    if chat_id in tag_data:
        step = tag_data[chat_id]['step']
        
        if step < len(TAG_QUESTIONS):
            kolom_db = TAG_QUESTIONS[step][0]
            jawaban = body if body != '-' else ''
            
            if kolom_db == 'tgl_aktif' and jawaban != '':
                try:
                    valid_date = datetime.strptime(jawaban, '%d/%m/%Y')
                    jawaban = valid_date.strftime('%Y-%m-%d')
                except ValueError:
                    send_wa_bot(chat_id, "❌ *Format Salah!* Gunakan format DD/MM/YYYY. Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            if kolom_db == 'tgl_jatuh_tempo' and jawaban != '':
                if not jawaban.isdigit() or not (1 <= int(jawaban) <= 31):
                    send_wa_bot(chat_id, "❌ *Format Salah!* Masukkan hanya Angka Tanggal (1-31). Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            if kolom_db == 'biaya_paket' and jawaban != '':
                if not jawaban.isdigit():
                    send_wa_bot(chat_id, "❌ *Format Salah!* Masukkan nominal uang hanya Angka. Masukkan ulang:")
                    return jsonify({"status": "success"}), 200
                    
            if kolom_db == 'no_hp' and jawaban != '':
                if not jawaban.isdigit():
                    send_wa_bot(chat_id, "❌ *Format Salah!* No HP hanya boleh Angka. Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            tag_data[chat_id]['data'][kolom_db] = jawaban
            step += 1
            tag_data[chat_id]['step'] = step

            if step < len(TAG_QUESTIONS):
                send_wa_bot(chat_id, f"Masukkan *{TAG_QUESTIONS[step][1]}*:")
            else:
                data_tagihan = tag_data[chat_id]['data']
                no_hp = data_tagihan['no_hp']
                if no_hp.startswith('0'):
                    no_hp = '62' + no_hp[1:]
                    
                biaya_prorata, jumlah_hari = hitung_biaya_prorata(data_tagihan['tgl_aktif'], data_tagihan['tgl_jatuh_tempo'], data_tagihan['biaya_paket'])
                
                try:
                    prorata_format = f"{int(biaya_prorata):,}".replace(",", ".")
                    normal_format = f"{int(data_tagihan['biaya_paket']):,}".replace(",", ".")
                except Exception:
                    prorata_format = biaya_prorata
                    normal_format = data_tagihan['biaya_paket']

                pesan_pelanggan = (
                    f"Halo Bapak/Ibu Pelanggan Setia,\n\n"
                    f"Kami informasikan bahwa layanan internet Anda saat ini telah memasuki *Bulan Ke-2*.\n\n"
                    f"📌 *Rincian Tagihan Bulan Ini*:\n"
                    f"• Tanggal Jatuh Tempo: Tanggal *{data_tagihan['tgl_jatuh_tempo']}*\n"
                    f"• Tagihan Prorata ({jumlah_hari} Hari): *Rp {prorata_format}*\n"
                    f"  _(Dihitung proporsional dari hari ke-2 aktif hingga tanggal jatuh tempo terdekat)_\n\n"
                    f"Link Pembayaran : https://ibb.co.com/zTQNjd4c\n\n"
                    f"Mohon kirim bukti pembayaran ke nomor ini, lakukan pembayaran tepat waktu sebelum tanggal jatuh tempo untuk menghindari isolir sistem otomatis. Terima kasih atas kepercayaan Anda! 🙏"
                )
                
                payload = {'target': no_hp, 'message': pesan_pelanggan}
                headers = {'Authorization': WA_API_TOKEN}
                
                try:
                    resp = requests.post(WA_API_PELANGGAN, json=payload, headers=headers, timeout=5)
                    if resp.status_code == 200:
                        send_wa_bot(chat_id, f"✅ Pesan penagihan bulan ke-2 (Rp {prorata_format}) sukses dikirim ke pelanggan ({no_hp}).")
                    else:
                        send_wa_bot(chat_id, f"⚠️ Pesan tagihan terkirim ke API, namun merespon dengan status code {resp.status_code}.")
                except Exception as e:
                    send_wa_bot(chat_id, f"❌ Gagal mengirim pesan ke pelanggan via port 8001: {str(e)}")
                    
                del tag_data[chat_id]
                
        return jsonify({"status": "success"}), 200

    if body.lower() in ['/psb', 'psb'] and chat_id not in user_data and chat_id not in tag_data:
        if not os.path.exists(TEMPLATE_MASTER):
            send_wa_bot(chat_id, "⚠️ ERROR: File master template-import-pelanggan.xlsx tidak ditemukan di NAS!")
            return jsonify({"status": "error"}), 200
            
        user_data[chat_id] = {'step': 0, 'data': {}}
        pesan = f"📝 *[PSB] Mulai Pendataan Pasang Baru*\n_Ketik batal untuk membatalkan._\n\nMasukkan *{QUESTIONS[0][1]}*:"
        send_wa_bot(chat_id, pesan)
        return jsonify({"status": "success"}), 200

    if chat_id in user_data:
        step = user_data[chat_id]['step']
        
        if step < len(QUESTIONS):
            kolom_db = QUESTIONS[step][0]
            jawaban = body if body != '-' else ''
            
            if kolom_db == 'nik*' and jawaban != '':
                if not jawaban.isdigit() or len(jawaban) != 16:
                    send_wa_bot(chat_id, "❌ *Format Salah!* NIK wajib 16 digit Angka. Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            if kolom_db == 'no_hp*' and jawaban != '':
                if not jawaban.isdigit():
                    send_wa_bot(chat_id, "❌ *Format Salah!* No HP hanya boleh Angka. Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            if kolom_db == 'tgl_aktif*' and jawaban != '':
                try:
                    valid_date = datetime.strptime(jawaban, '%d/%m/%Y')
                    jawaban = valid_date.strftime('%Y-%m-%d')
                except ValueError:
                    send_wa_bot(chat_id, "❌ *Format Salah!* Gunakan format DD/MM/YYYY. Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            if kolom_db == 'tgl_jatuh_tempo' and jawaban != '':
                if not jawaban.isdigit() or not (1 <= int(jawaban) <= 31):
                    send_wa_bot(chat_id, "❌ *Format Salah!* Masukkan hanya Angka Tanggal (1-31). Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            if kolom_db == 'biaya_paket' and jawaban != '':
                if not jawaban.isdigit():
                    send_wa_bot(chat_id, "❌ *Format Salah!* Masukkan nominal uang hanya Angka. Masukkan ulang:")
                    return jsonify({"status": "success"}), 200

            user_data[chat_id]['data'][kolom_db] = jawaban
            step += 1
            user_data[chat_id]['step'] = step

            if step < len(QUESTIONS):
                send_wa_bot(chat_id, f"Masukkan *{QUESTIONS[step][1]}*:")
            else:
                send_wa_bot(chat_id, "📍 Kirimkan *Lokasi (Share Location)* pelanggan:\n_ (Ketik '-' untuk lewati)_")
            return jsonify({"status": "success"}), 200

        elif step == len(QUESTIONS):
            if msg_type == 'location' or (payload.get('lat') and payload.get('lng')):
                user_data[chat_id]['data']['lat*'] = payload.get('lat')
                user_data[chat_id]['data']['lon*'] = payload.get('lng')
            elif body == '-':
                user_data[chat_id]['data']['lat*'] = ''
                user_data[chat_id]['data']['lon*'] = ''
            else:
                send_wa_bot(chat_id, "⚠️ Harap gunakan fitur *Share Location* atau ketik '-'. Coba lagi:")
                return jsonify({"status": "success"}), 200
            
            user_data[chat_id]['step'] = 9
            send_wa_bot(chat_id, "🪪 Kirimkan *Foto KTP* (⚠️ Wajib Gambar/Foto):")
            return jsonify({"status": "success"}), 200

        elif step == 9:
            if msg_type != 'image' and not payload.get('media_base64'):
                send_wa_bot(chat_id, "❌ Wajib mengirimkan gambar Foto KTP. Silakan kirim ulang:")
                return jsonify({"status": "success"}), 200
            
            nama_pelanggan = user_data[chat_id]['data'].get('nama_pelanggan*', 'TanpaNama').replace(" ", "_").replace("/", "_")
            filepath = os.path.join(KTP_PATH, f"KTP_{nama_pelanggan}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg")
            
            with open(filepath, "wb") as fh:
                fh.write(base64.b64decode(payload.get('media_base64')))
            
            beri_watermark_foto(filepath, user_data[chat_id]['data'].get('lat*', '-'), user_data[chat_id]['data'].get('lon*', '-'))
            user_data[chat_id]['final_ktp_path'] = filepath
            
            user_data[chat_id]['step'] = 10
            send_wa_bot(chat_id, "🏠 Terakhir, Kirimkan *Foto Rumah* (⚠️ Wajib Gambar/Foto):")
            return jsonify({"status": "success"}), 200

        elif step == 10:
            if msg_type != 'image' and not payload.get('media_base64'):
                send_wa_bot(chat_id, "❌ Wajib mengirimkan gambar Foto Rumah. Silakan kirim ulang:")
                return jsonify({"status": "success"}), 200
            
            nama_pelanggan = user_data[chat_id]['data'].get('nama_pelanggan*', 'TanpaNama').replace(" ", "_").replace("/", "_")
            filepath = os.path.join(RUMAH_PATH, f"RUMAH_{nama_pelanggan}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg")
            
            with open(filepath, "wb") as fh:
                fh.write(base64.b64decode(payload.get('media_base64')))
                
            beri_watermark_foto(filepath, user_data[chat_id]['data'].get('lat*', '-'), user_data[chat_id]['data'].get('lon*', '-'))
            user_data[chat_id]['final_rumah_path'] = filepath
            
            send_wa_bot(chat_id, "⏳ Semua data lengkap! Sedang memproses penyimpanan file ke server NAS...")
            simpan_data_keseluruhan(chat_id)
            return jsonify({"status": "success"}), 200

    return jsonify({"status": "idle"}), 200

if __name__ == '__main__':
    # Diubah menjadi port 5555 sesuai permintaan
    print("Sistem Python Bot WA PSB Ready Port 5555 (DOCKER MODE)...")
    app.run(host='0.0.0.0', port=5555, debug=False)
